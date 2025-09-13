import os
import asyncio
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import date, datetime, timedelta
from app.crud.attendance import get_monthly_attendance_report

def smart_column_width(worksheet, min_width=8, max_width=50, padding=2):
    """
    Ustunlar kengligini mazmuniga qarab aqlli sozlash
    
    Args:
        worksheet: Excel worksheet objekti
        min_width: Minimal ustun kengligi
        max_width: Maksimal ustun kengligi  
        padding: Qo'shimcha joy (kenglik)
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Optimal kenglikni hisoblash
        adjusted_width = min(max(max_length + padding, min_width), max_width)
        
        # Maxsus ustunlar uchun qoidalar
        header_value = str(column[0].value).lower() if column[0].value else ""
        
        if any(word in header_value for word in ["ism", "name"]):
            adjusted_width = max(adjusted_width, 25)
        elif any(word in header_value for word in ["lavozim", "position"]):
            adjusted_width = max(adjusted_width, 20)
        elif any(word in header_value for word in ["izoh", "comment"]):
            adjusted_width = max(adjusted_width, 35)
        elif any(word in header_value for word in ["maosh", "salary"]):
            adjusted_width = max(adjusted_width, 18)
        elif any(word in header_value for word in ["vaqt", "time"]):
            adjusted_width = max(adjusted_width, 15)
        elif header_value == "№":
            adjusted_width = 6
        elif "%" in header_value:
            adjusted_width = max(adjusted_width, 12)
            
        worksheet.column_dimensions[column_letter].width = adjusted_width

async def generate_detailed_monthly_report(db: AsyncSession, year: int, month: int):
    """Batafsil oylik Excel hisobotini yaratish - vaqt va maosh ma'lumotlari bilan"""
    
    # Ma'lumotlarni olish
    report_data = await get_monthly_attendance_report(db, month, year)
    
    # Excel faylini yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d} Batafsil Hisoboti"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Header - batafsil ma'lumotlar
    headers = [
        "№", "Xodim ismi", "Lavozimi", "Jami kelgan kunlar", "Kechikkanlar", 
        "Erta ketganlar", "Asosiy maosh", "Yakuniy maosh", "Chegirmalar", "Izoh"
    ]
    
    # Header qo'shish
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Ma'lumotlarni qo'shish
    for row_idx, row_data in enumerate(report_data, 2):
        # Salary ma'lumotlari
        salary_info = row_data.get("salary_info", {})
        base_salary = salary_info.get("base_salary", 0)
        final_salary = salary_info.get("final_salary", 0)
        total_deductions = salary_info.get("total_deductions", 0)
        
        # Izoh yaratish
        izoh_parts = []
        if row_data["late_days"] > 0:
            izoh_parts.append(f"Kechikish: {row_data['late_days']} kun")
        if row_data.get("early_departure_days", 0) > 0:
            izoh_parts.append(f"Erta ketish: {row_data.get('early_departure_days', 0)} kun")
        if total_deductions > 0:
            izoh_parts.append(f"Jarim: {total_deductions:,.0f} so'm")
        izoh = "; ".join(izoh_parts) if izoh_parts else "Yaxshi ish"
        
        row_values = [
            row_idx - 1,  # №
            row_data["employee_name"],
            row_data["position"] if row_data["position"] else "Belgilanmagan",
            row_data["present_days"],
            row_data["late_days"],
            row_data.get("early_departure_days", 0),
            f"{base_salary:,.0f}" if base_salary > 0 else "Belgilanmagan",
            f"{final_salary:,.0f}" if final_salary > 0 else "0",
            f"{total_deductions:,.0f}" if total_deductions > 0 else "0",
            izoh
        ]
        
        for col, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = center_alignment
            cell.border = border
            
            # Rangli format
            if col == 5 and isinstance(value, int) and value > 0:  # Kechikkanlar
                cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
            if col == 6 and isinstance(value, int) and value > 0:  # Erta ketganlar
                cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
            if col == 9 and isinstance(value, str) and "0" not in value:  # Chegirmalar
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Ma'lumotlarni qo'shish
    for row_idx, row_data in enumerate(report_data, 2):
        # Salary ma'lumotlari
        salary_info = row_data.get("salary_info", {})
        base_salary = salary_info.get("base_salary", 0)
        final_salary = salary_info.get("final_salary", 0)
        total_deductions = salary_info.get("total_deductions", 0)
        
        # Izoh yaratish
        izoh_parts = []
        if row_data["late_days"] > 0:
            izoh_parts.append(f"Kechikish: {row_data['late_days']} kun")
        if row_data.get("early_departure_days", 0) > 0:
            izoh_parts.append(f"Erta ketish: {row_data.get('early_departure_days', 0)} kun")
        if total_deductions > 0:
            izoh_parts.append(f"Jarim: {total_deductions:,.0f} so'm")
        izoh = "; ".join(izoh_parts) if izoh_parts else "Yaxshi ish"
        
        row_values = [
            row_idx - 1,  # №
            row_data["employee_name"],
            row_data["position"] if row_data["position"] else "Belgilanmagan",
            row_data["present_days"],
            row_data["late_days"],
            row_data.get("early_departure_days", 0),
            f"{base_salary:,.0f}" if base_salary > 0 else "Belgilanmagan",
            f"{final_salary:,.0f}" if final_salary > 0 else "0",
            f"{total_deductions:,.0f}" if total_deductions > 0 else "0",
            izoh
        ]
        
        for col, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = center_alignment
            cell.border = border
            
            # Rangli format
            if col == 5 and isinstance(value, int) and value > 0:  # Kechikkanlar
                cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
            if col == 6 and isinstance(value, int) and value > 0:  # Erta ketganlar
                cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
            if col == 9 and isinstance(value, str) and "0" not in value:  # Chegirmalar
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Aqlli ustun kengligini sozlash
    smart_column_width(ws)
    
    # Fayl nomini yaratish
    os.makedirs("reports", exist_ok=True)
    file_path = f"reports/batafsil_hisoboti_{year}_{month:02d}.xlsx"
    wb.save(file_path)
    
    return file_path

async def generate_monthly_report(db: AsyncSession, year: int, month: int):
    """Oylik Excel hisobotini yaratish - yangilangan versiya maoshlar bilan"""
    
    # Ma'lumotlarni olish
    report_data = await get_monthly_attendance_report(db, month, year)
    
    # Oyning umumiy ish kunlari (dam olish kunlarisiz)
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    working_days = 0
    current_date = start_date
    while current_date <= end_date:
        # Faqat yakshanba dam olish kuni (Sunday=6), qolgan kunlar ish kunlari
        if current_date.weekday() != 6:  # Sunday=6, qolgan kunlar ish kuni
            working_days += 1
        current_date += timedelta(days=1)
    
    # Excel faylini yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d} Davomat Hisoboti"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Header - davomat ma'lumotlari
    headers = [
        "№", "Xodim ismi", "Lavozimi", "Umumiy ish kunlari", "Jami kelgan kunlar", 
        "Kechikkan kunlar", "Erta ketgan kunlar", "Ishga kelmagan kunlar", "Davomat %"
    ]
    
    # Header qo'shish
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Ma'lumotlarni qo'shish
    for row_idx, row_data in enumerate(report_data, 2):
        attendance_percentage = (row_data["present_days"] / working_days * 100) if working_days > 0 else 0
        
        row_values = [
            row_idx - 1,  # №
            row_data["employee_name"],
            row_data["position"] if row_data["position"] else "Belgilanmagan",
            working_days,
            row_data["present_days"],
            row_data["late_days"],
            row_data.get("early_departure_days", 0),
            working_days - row_data["present_days"],  # Yo'qliklar
            f"{attendance_percentage:.1f}%"
        ]
        
        for col, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = center_alignment
            cell.border = border
            
            # Muammolar uchun qizil rang
            try:
                numeric_value = float(str(value).replace(",", "").replace("%", "")) if isinstance(value, str) else value
                if col == 6 and numeric_value > 0:  # Kechikkanlar ustuni
                    cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
                if col == 7 and numeric_value > 0:  # Erta ketganlar ustuni  
                    cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
                if col == 8 and numeric_value > 0:  # Yo'qliklar ustuni
                    cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
            except (ValueError, TypeError):
                pass  # Agar raqamga aylantira olmasak, rangni o'zgartirmaymiz
    
    # Aqlli ustun kengligini sozlash
    smart_column_width(ws)
    
    # Jami statistika qo'shish
    total_row = len(report_data) + 3
    ws.cell(row=total_row, column=1, value="JAMI:").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(row["present_days"] for row in report_data)).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=sum(row["late_days"] for row in report_data)).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=sum(row.get("early_departure_days", 0) for row in report_data)).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=sum(working_days - row["present_days"] for row in report_data)).font = Font(bold=True)
    
    # Fayl nomini yaratish
    os.makedirs("reports", exist_ok=True)
    file_path = f"reports/davomat_hisoboti_{year}_{month:02d}.xlsx"
    wb.save(file_path)
    
    return file_path

async def delete_file_after_delay(file_path: str, delay_seconds: int = 300):
    """Faylni ma'lum vaqtdan keyin o'chirish (default: 5 daqiqa)"""
    await asyncio.sleep(delay_seconds)
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"🗑️ Fayl o'chirildi: {file_path}")
    except Exception as e:
        print(f"❌ Faylni o'chirishda xato: {e}")

async def generate_daily_report(db: AsyncSession, target_date: date):
    """Kunlik batafsil hisobotni yaratish - vaqt ma'lumotlari bilan"""
    from app.crud.attendance import get_daily_attendance
    
    attendances = await get_daily_attendance(db, target_date)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{target_date} Kunlik Davomat"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Headers
    headers = [
        "№", "Xodim ismi", "Lavozimi", "Harakat", "Vaqti", 
        "Kechikdi", "Vaqt farqi", "Izoh"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Ma'lumotlar
    for row_idx, attendance in enumerate(attendances, 2):
        # Employee ma'lumotlarini olish
        from app.crud.employee import get_employee_by_id
        employee = await get_employee_by_id(db, attendance.employee_id)
        
        # Vaqt farqi va izoh hisoblash
        vaqt_farqi = ""
        izoh = ""
        
        if attendance.check_type.value == "IN":
            if attendance.is_late:
                # Kechikish vaqtini hisoblash (9:30 dan necha daqiqa kech)
                from datetime import time, datetime, timedelta
                work_start = time(9, 30)
                work_start_datetime = datetime.combine(attendance.check_time.date(), work_start)
                if attendance.check_time > work_start_datetime:
                    late_minutes = (attendance.check_time - work_start_datetime).seconds // 60
                    vaqt_farqi = f"{late_minutes} daqiqa kech"
                    izoh = f"Kechikish: {late_minutes} daqiqa"
                else:
                    vaqt_farqi = "Vaqtida"
                    izoh = "Vaqtida keldi"
            else:
                vaqt_farqi = "Vaqtida"
                izoh = "Vaqtida keldi"
        
        elif attendance.check_type.value == "OUT":
            vaqt_farqi = "Vaqtida"
            izoh = "Normal vaqtda ketdi"
        
        row_values = [
            row_idx - 1,
            employee.full_name if employee else "Noma'lum",
            employee.position if employee and employee.position else "N/A",
            "Keldi" if attendance.check_type.value == "IN" else "Ketdi",
            attendance.check_time.strftime("%H:%M:%S"),
            "Ha" if attendance.is_late else "Yo'q",
            vaqt_farqi,
            izoh
        ]
        
        for col, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = center_alignment
            cell.border = border
            
            # Rangli format
            if col == 6 and value == "Ha":  # Kechikdi
                cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
            if col == 7 and value == "Ha":  # Erta ketdi
                cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
    
    # Aqlli ustun kengligini sozlash
    smart_column_width(ws)
    
    # Fayl saqlash
    os.makedirs("reports", exist_ok=True)
    file_path = f"reports/kunlik_batafsil_{target_date}.xlsx"
    wb.save(file_path)
    
    return file_path
